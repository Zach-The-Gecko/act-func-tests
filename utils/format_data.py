import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Define paths
csv_path = Path(__file__).parent / "output.csv"
json_path = Path(__file__).parent.parent / "experiment_results.json"
output_dir = Path(__file__).parent.parent / "formatted_data"

# Create output directory
output_dir.mkdir(exist_ok=True)

# Load JSON data
with open(json_path, 'r') as f:
    experiments = json.load(f)

# Read CSV data (skip metadata line)
csv_data = []
with open(csv_path, 'r') as f:
    next(f)  # Skip metadata line
    reader = csv.DictReader(f)
    for row in reader:
        csv_data.append(row)

date_str = "2026-02-20"
parsed_csv_rows = []

for row in csv_data:
    time_str = (row.get('Time') or '').strip()
    gpu_value_str = (row.get('GPU Package') or '').strip()

    if not time_str or not gpu_value_str:
        continue

    datetime_str = f"{date_str} {time_str}"
    try:
        dt = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
        gpu_value = float(gpu_value_str)
        parsed_csv_rows.append((dt, gpu_value))
    except ValueError:
        continue

# Process each activation function and collect all runs
for activation_func, data in experiments.items():
    # Dictionary to store all run data by index
    all_runs = {}
    max_data_points = 0
    run_metadata = []

    for run in data['runs']:
        run_num = run['run_number']
        start_time = datetime.strptime(run['start'], "%Y-%m-%d %H:%M:%S.%f")
        end_time = datetime.strptime(run['end'], "%Y-%m-%d %H:%M:%S.%f")

        # Filter CSV data for this run
        run_data = []
        for timestamp, gpu_value in parsed_csv_rows:
            if start_time <= timestamp <= end_time:
                run_data.append(gpu_value)

        all_runs[run_num] = run_data
        max_data_points = max(max_data_points, len(run_data))
        run_metadata.append((run_num, len(run_data)))
        print(
            f"Processed {activation_func} run {run_num}: {len(run_data)} data points")

    # Create Excel workbook with separate sheets for summary and data
    excel_file = output_dir / f"{activation_func}.xlsx"
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(['Run', 'Start Time', 'End Time',
                      'Duration (s)', 'Data Points'])

    # Style header row
    header_fill = PatternFill(start_color="4472C4",
                              end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    for run in data['runs']:
        run_num = run['run_number']
        start_time = datetime.strptime(
            run['start'], "%Y-%m-%d %H:%M:%S.%f")
        end_time = datetime.strptime(run['end'], "%Y-%m-%d %H:%M:%S.%f")

        count = 0
        for timestamp, _ in parsed_csv_rows:
            if start_time <= timestamp <= end_time:
                count += 1

        ws_summary.append([
            run_num,
            run['start'],
            run['end'],
            round(run['duration_seconds'], 2),
            count
        ])

    # Adjust column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 30
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15

    # Create Data sheet
    ws_data = wb.create_sheet("GPU Power Data")

    # Write header
    header = ['Row']
    for run_num in sorted(all_runs.keys()):
        header.append(f'Run {run_num}')
    ws_data.append(header)

    # Style header row
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_idx in range(max_data_points):
        row = [row_idx + 1]
        for run_num in sorted(all_runs.keys()):
            if row_idx < len(all_runs[run_num]):
                row.append(all_runs[run_num][row_idx])
            else:
                row.append('')
        ws_data.append(row)

    # Adjust column widths for data sheet
    ws_data.column_dimensions['A'].width = 8
    for col_idx in range(2, len(all_runs) + 2):
        ws_data.column_dimensions[chr(64 + col_idx)].width = 15

    # Save workbook
    wb.save(excel_file)
    print(f"Created Excel file for {activation_func}")

print(f"\nFormatted data saved to: {output_dir}")
